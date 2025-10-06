from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Customer, Payment
from .serializers import CustomerSerializer, PaymentSerializer
from rest_framework.permissions import AllowAny
from rest_framework.decorators import permission_classes

# List & Create Customers
@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def customers_list_create(request):
    if request.method == 'GET':
        q = request.query_params.get('q', '')
        customers = Customer.objects.filter(name__icontains=q) | Customer.objects.filter(village__icontains=q) | Customer.objects.filter(phone__icontains=q)
        serializer = CustomerSerializer(customers, many=True)
        return Response({'items': serializer.data})

    elif request.method == 'POST':
        serializer = CustomerSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'item': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Retrieve, Update, Delete Customer
@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([AllowAny])
def customer_detail(request, pk):
    try:
        customer = Customer.objects.get(pk=pk)
    except Customer.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = CustomerSerializer(customer)
        return Response({'item': serializer.data})

    elif request.method == 'PATCH':
        serializer = CustomerSerializer(customer, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'item': serializer.data})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        customer.delete()
        return Response({'ok': True})


# Add Payment
@api_view(['POST'])
@permission_classes([AllowAny])
def add_payment(request, pk):
    try:
        customer = Customer.objects.get(pk=pk)
    except Customer.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    data = request.data
    amount = data.get('amount')
    date = data.get('date')

    if amount is None or amount <= 0 or not date:
        return Response({'error': 'amount>0 and date required'}, status=status.HTTP_400_BAD_REQUEST)

    payment = Payment.objects.create(customer=customer, amount=amount, date=date)
    serializer = CustomerSerializer(customer)
    return Response({'item': serializer.data})
import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
@api_view(['GET'])
@permission_classes([AllowAny])
def export_customers_excel(request):
    customers = Customer.objects.all()
    serializer = CustomerSerializer(customers, many=True)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Define headers
    headers = ['ID', 'Name', 'Village', 'Phone', 'Total Amount', 'Paid', 'Due']
    ws.append(headers)

    # Add data rows
    for cust in serializer.data:
        ws.append([
            cust['id'],
            cust['name'],
            cust['village'],
            cust['phone'],
            cust['total_amount'],
            cust['paid'],
            cust['due'],
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Prepare HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customers.xlsx'
    wb.save(response)
    return response
